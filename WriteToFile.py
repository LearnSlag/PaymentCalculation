# -*- coding: UTF-8 -*-
import os
import openpyxl
import csv

def write_statistics_data(sheet_name, list_subproject, list_year_month, table_data):
    work_book = openpyxl.load_workbook('腾讯云费用统计数据.xlsx')
    work_sheets = work_book.sheetnames
    work_sheet = list(filter(lambda x: sheet_name in x, work_sheets))[0]
    work_book.remove_sheet(work_book.get_sheet_by_name(work_sheet))
    sheet_statistics = work_book.create_sheet(work_sheet) 
    #sheet_statistics = work_book[work_sheet]
    sheet_statistics.alignment= openpyxl.styles.Alignment(horizontal='general',
                    vertical='bottom',
                    text_rotation=0,
                    wrap_text=False,
                    shrink_to_fit=False,
                    indent=0)
    sheet_statistics.cell(row=3, column=1, value = '求和项:现金账户支出(元)')
    sheet_statistics.cell(row=3, column=2, value = '使用结束时间（月）')
    sheet_statistics.cell(row=4, column=1, value = '子产品名称')
    for subproject in list_subproject:
        sheet_statistics.cell(row=list_subproject.index(subproject)+5, column=1, value = subproject)
        for year_month in list_year_month:
            sheet_statistics.cell(row=4, column=list_year_month.index(year_month)+2, value = year_month)
            if subproject in table_data:
                if year_month in table_data[subproject]:
                    sheet_statistics.cell(row=list_subproject.index(subproject)+5, column=list_year_month.index(year_month)+2, value = table_data[subproject][year_month])

    work_book.save('腾讯云费用统计数据.xlsx')



