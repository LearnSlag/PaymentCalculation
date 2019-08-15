# -*- coding: UTF-8 -*-
import os
import re
import csv
import openpyxl
import datetime
import WriteToFile

def read_preparative_data(file_name, summarize_file_name):
    if os.path.exists(file_name):
        with open(summarize_file_name,'a', encoding = 'utf-8',newline='') as sf:
            with open(file_name, 'r', encoding = 'utf-8') as f:
                list_data = []
                list_title = {}
                preparative_csv_file = csv.reader(f)
                list_data_temp = list(preparative_csv_file)
                list_title = list_data_temp[0]
                row_temp = 0
                for row in list_data_temp[1:]:
                    csv.writer(sf).writerow(row)#将2.csv内容追加到1.csv的后面
                    dict_used = {}
                    dict_among = {}
                    column_temp = 0
                    for data in row:
                        dict_among[list_title[column_temp]] = data
                        column_temp = column_temp + 1
                #print(dict_among)
                    dict_used['子产品名称'] = dict_among['子产品名称']
                    dict_used['现金账户支出(元)'] = dict_among['现金账户支出(元)']
                    dict_used['结束使用时间'] = datetime.datetime.strptime(dict_among['结束使用时间'][0:16], "%Y/%m/%d %H:%M").strftime("%Y-%m") 
                    list_data.append(dict_used)
                    row_temp = row_temp + 1
            return list_data
    else:
        return None


def read_statistics_data(sheet_name):
    work_book = openpyxl.load_workbook('腾讯云费用统计数据.xlsx')
    work_sheets = work_book.sheetnames
    #statistics_data_official = list(filter(lambda x: '正式账户' in x, work_sheets)).pop()
    statistics_data = list(filter(lambda x: sheet_name in x, work_sheets)).pop()
    sheet_statistics = work_book[statistics_data]
    list_subproject = []
    list_year_month = []
    dict_statistics_data= {}
    if sheet_statistics.cell(row=3, column=1).value is not None:
        list_statistics = list(sheet_statistics)
        for row_among in list_statistics:
            dict_among = {}
            for cell_among in list(row_among):
                if cell_among.row is 4 and cell_among.column > 1:
                    list_year_month.append(cell_among.value)
                    continue
                else:
                    if cell_among.column is 1 and cell_among.row > 4:
                        list_subproject.append(cell_among.value)
                        continue
                    else:
                        if cell_among.value is not None and cell_among.row > 4 and cell_among.column > 1:
                            dict_among[list_year_month[cell_among.column - 2]] = cell_among.value
                if cell_among.row > 4 and cell_among.column > 1:
                    dict_statistics_data[list_subproject[cell_among.row - 5]] = dict_among
    work_book.save('腾讯云费用统计数据.xlsx')
    #print(list_subproject)
    #print(list_year_month)
    #print(dict_statistics_data)
    return list_subproject, list_year_month, dict_statistics_data



